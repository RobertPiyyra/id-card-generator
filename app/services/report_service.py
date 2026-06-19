"""
Report Generation Service
Generates PDF and Excel reports for various data types.
Isolated module - reads from existing models only.
"""
import io
import logging
from datetime import datetime, timezone, timedelta

from models import db, Student, Template, ActivityLog, BulkJob

logger = logging.getLogger(__name__)


def generate_student_report(format: str = 'xlsx', school_name: str = None,
                            template_id: int = None) -> io.BytesIO:
    """
    Generate a student report in XLSX or CSV format.
    Returns a BytesIO buffer.
    """
    query = Student.query
    if school_name:
        query = query.filter_by(school_name=school_name)
    if template_id:
        query = query.filter_by(template_id=template_id)

    students = query.order_by(Student.created_at.desc()).all()

    headers = ['ID', 'Name', 'Father Name', 'Class', 'DOB', 'School', 'Phone',
               'Address', 'Email', 'Has Photo', 'Card Generated', 'Created At']

    rows = []
    for s in students:
        rows.append([
            s.id,
            s.name or '',
            s.father_name or '',
            s.class_name or '',
            s.dob or '',
            s.school_name or '',
            s.phone or '',
            s.address or '',
            s.email or '',
            'Yes' if (s.photo_url or s.photo_filename) else 'No',
            'Yes' if s.image_url else 'No',
            s.created_at.strftime('%Y-%m-%d %H:%M') if s.created_at else '',
        ])

    if format == 'csv':
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        result = io.BytesIO(buf.getvalue().encode('utf-8'))
        result.name = 'student_report.csv'
        return result

    # Default XLSX
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Students'

        # Header styling
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'student_report.xlsx'
        return buf

    except ImportError:
        # Fallback to CSV if openpyxl not available
        logger.warning("openpyxl not installed, falling back to CSV")
        return generate_student_report(format='csv', school_name=school_name, template_id=template_id)


def generate_activity_report(format: str = 'xlsx', days: int = 30) -> io.BytesIO:
    """Generate an activity log report."""
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    activities = ActivityLog.query.filter(
        ActivityLog.timestamp >= cutoff
    ).order_by(ActivityLog.timestamp.desc()).all()

    headers = ['Timestamp', 'Actor', 'Action', 'Target', 'Details', 'IP Address']
    rows = []
    for a in activities:
        rows.append([
            a.timestamp.strftime('%Y-%m-%d %H:%M:%S') if a.timestamp else '',
            a.actor or '',
            a.action or '',
            a.target or '',
            a.details or '',
            a.ip_address or '',
        ])

    if format == 'csv':
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        result = io.BytesIO(buf.getvalue().encode('utf-8'))
        result.name = 'activity_report.csv'
        return result

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Activity Log'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'activity_report.xlsx'
        return buf
    except ImportError:
        return generate_activity_report(format='csv', days=days)


def generate_bulk_job_report(format: str = 'xlsx') -> io.BytesIO:
    """Generate a bulk job execution report."""
    jobs = BulkJob.query.order_by(BulkJob.created_at.desc()).limit(100).all()

    headers = ['ID', 'Template ID', 'Type', 'Status', 'Total Items', 'Processed',
               'Failed', 'Created By', 'Created At']
    rows = []
    for j in jobs:
        rows.append([
            j.id, j.template_id or '', j.job_type or '', j.status or '',
            j.total_items, j.processed_items, j.failed_items,
            j.created_by or '',
            j.created_at.strftime('%Y-%m-%d %H:%M') if j.created_at else '',
        ])

    if format == 'csv':
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        result = io.BytesIO(buf.getvalue().encode('utf-8'))
        result.name = 'bulk_job_report.csv'
        return result

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bulk Jobs'
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'bulk_job_report.xlsx'
        return buf
    except ImportError:
        return generate_bulk_job_report(format='csv')
